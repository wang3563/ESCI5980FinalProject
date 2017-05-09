#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Tue Apr 25 09:29:17 2017

@author: Julia Nissen, Zongyi Wang
"""
import sys
import Tkinter as tk
import tkFileDialog as filedialog
import tkMessageBox as messagebox
import numpy as np
from openpyxl import load_workbook
from scipy.optimize import fsolve

class Application(tk.Frame):
    '''
    GUI for age caculation for Larry Edwards trace metal lab
    created by Julia and Nick

    May 2017
    '''
    def __init__(self, master):
        tk.Frame.__init__(self,master)
        self.dialog_frame = tk.Frame(self)
        self.dialog_frame.pack(padx = 20, pady = 15, anchor = 'w')
        tk.Label(self.dialog_frame, text = "Welcome to age calculation!" ).grid(row = 0, column = 0, sticky = 'e')
        self.master.title("Age Calculation")
        self.create_widgets()
        self.pack()
        
    def create_widgets(self):
        # some entry widgets
        tk.Label(self.dialog_frame, text = "Enter spike information(choose from: DIII-B, DIII-A, 1I, 1H)  ").grid(row = 1, column = 0, sticky = 'w')
        self.spikeinput = tk.Entry(self.dialog_frame, background = 'white', width = 24)
        self.spikeinput.grid(row = 1, column = 1, sticky = 'w')
        self.spikeinput.focus_set()
        tk.Label(self.dialog_frame, text = "Enter abundant sensitivity for 237U-238U ").grid(row = 2, column = 0, sticky = 'w')
        self.AS1_input = tk.Entry(self.dialog_frame, background = 'white', width = 24)
        self.AS1_input.grid(row =2, column = 1, sticky = 'w')
        self.AS1_input.focus_set()

        tk.Label(self.dialog_frame, text = "Enter sample weight(g) ").grid(row = 3, column = 0, sticky = 'w')
        self.samplewt = tk.Entry(self.dialog_frame, background = 'white', width = 24)
        self.samplewt.grid(row =3, column = 1, sticky = 'w')
        self.samplewt.focus_set()
        
        
        tk.Label(self.dialog_frame, text = "Enter spike weight(g) ").grid(row = 4, column = 0, sticky = 'w')
        self.spikewt = tk.Entry(self.dialog_frame, background = 'white', width = 24)
        self.spikewt.grid(row =4, column = 1, sticky = 'w')
        self.spikewt.focus_set()
        
        tk.Label(self.dialog_frame, text = "Enter chem spike weight(g) ").grid(row = 5, column = 0, sticky = 'w')
        self.chemspikewt = tk.Entry(self.dialog_frame, background = 'white', width = 24)
        self.chemspikewt.grid(row =5, column = 1, sticky = 'w')
        self.chemspikewt.focus_set()
        
        tk.Label(self.dialog_frame, text = "Enter sample ID: ").grid(row = 6, column = 0, sticky = 'w')
        self.samplename = tk.Entry(self.dialog_frame, background = 'white', width = 24)
        self.samplename.grid(row =6, column = 1, sticky = 'w')
        self.samplename.focus_set()
        
        
        tk.Label(self.dialog_frame, text = "Enter measurement year  ").grid(row = 7, column = 0, sticky = 'w')
        self.year_input = tk.Entry(self.dialog_frame, background = 'white', width = 24)
        self.year_input.grid(row =7, column = 1, sticky = 'w')
        self.year_input.focus_set()
        
        tk.Label(self.dialog_frame, text = "Enter the row number for the calculation results to written into(starting from 6) ").grid(row = 8, column = 0, sticky = 'w')
        self.row_input = tk.Entry(self.dialog_frame, background = 'white', width = 24)
        self.row_input.grid(row =8, column = 1, sticky = 'w')
        self.row_input.focus_set()
        
         
        #Define submit and cancel buttons
        button_frame = tk.Frame(self)
        button_frame.pack(padx=15, pady=(0, 15), anchor='e')
        

        self.submit_button = tk.Button(button_frame, text='Submit', default='active', command=self.click_submit)
        self.submit_button.pack(side='right')

        self.cancel_button = tk.Button(button_frame, text='Cancel', command=self.click_cancel)
        self.cancel_button.pack(side='right')
        
        #upload buttons 
    
        self.u_meas_upload = tk.Button(self)
        self.u_meas_upload["text"] = "Upload U measurement file"
        self.u_meas_upload["command"] = self.file_upload_u_meas
    
        #self.u_meas_upload.grid(row = 10, column = 1)
        self.u_meas_upload.pack()
           
        self.th_meas_upload = tk.Button(self)
        self.th_meas_upload["text"] = "Upload Th measurement file"
        self.th_meas_upload["command"] = self.file_upload_th_meas
        #self.th_meas_upload.grid(row = 10, column = 2)
        self.th_meas_upload.pack()
        
        
        self.u_wash_upload = tk.Button(self)
        self.u_wash_upload["text"] = "Upload U wash file"
        self.u_wash_upload["command"] = self.file_upload_u_wash
        #self.u_wash_upload.grid(row = 11, column = 1)
        self.u_wash_upload.pack()
        
        self.th_wash_upload = tk.Button(self)
        self.th_wash_upload["text"] = "Upload Th wash file"
        self.th_wash_upload["command"] = self.file_upload_th_wash
        #self.th_wash_upload.grid(row = 11, column =4  )
        self.th_wash_upload.pack()
        
        self.u_chemblank_upload = tk.Button(self)
        self.u_chemblank_upload["text"] = "Upload U chem blank file"
        self.u_chemblank_upload["command"] = self.file_upload_u_chemblank
        #self.u_chemblank_upload.grid(row = 12, column =1  )
        self.u_chemblank_upload.pack()
        
        self.th_chemblank_upload = tk.Button(self)
        self.th_chemblank_upload["text"] = "Upload Th chem blank file"
        self.th_chemblank_upload["command"] = self.file_upload_th_chemblank
        #self.th_chemblank_upload.grid(row = 12, column =2)
        self.th_chemblank_upload.pack()
        
        self.u_chemblankwash_upload = tk.Button(self)
        self.u_chemblankwash_upload["text"] = "Upload U chem blank wash file"
        self.u_chemblankwash_upload["command"] = self.file_upload_u_chemblankwash
        #self.u_chemblankwash_upload.grid(row = 13, column =0 )
        self.u_chemblankwash_upload.pack()
        
        self.th_chemblankwash_upload = tk.Button(self)
        self.th_chemblankwash_upload["text"] = "Upload Th chem blank wash file"
        self.th_chemblankwash_upload["command"] = self.file_upload_th_chemblankwash
        #self.u_chemblankwash_upload.grid(row = 13, column =2 )
        self.th_chemblankwash_upload.pack()
        
        self.file_export_upload = tk.Button(self)
        self.file_export_upload ["text"] = "Upload age export file"
        self.file_export_upload["command"] = self.file_upload_export
        #self.file_export_upload.grid(row= 14, column = 1)
        self.file_export_upload.pack( )
        
        #age claculation button
        self.agecalc = tk.Button(self)
        self.agecalc["text"] = "Calculate age and export data"
        self.agecalc["command"] = self.Age_Calculation

        #self.agecalc.grid(row = 14, column = 2)
        self.agecalc.pack()
        # quit button
        self.quit = tk.Button(self, text="QUIT", fg="red",command=root.destroy)
        self.quit.pack()
    
        
    def file_upload_u_meas(self):
        
        filename_raw = filedialog.askopenfilename(parent=self)
        self.filename_u_meas = filename_raw
        
        try:
            self.file_u_meas = load_workbook(str(filename_raw))
            messagebox.showinfo("Success!", "You have uploaded your U measurement file! " )
        except OSError as err:
            messagebox.showwarning("Error", str(err))
        except:
            messagebox.showerror("Unexpected error:", str(sys.exc_info()[:]))
       
    def file_upload_th_meas(self):
        
        filename_raw = filedialog.askopenfilename(parent=self)
        self.filename_th_meas = filename_raw
        try:
            self.file_u_meas = load_workbook(filename_raw)
            messagebox.showinfo("Success!", "You have uploaded your Th measurement file! " )
        except OSError as err:
            messagebox.showwarning("Error", str(err))
        except:
            messagebox.showerror("Unexpected error:", str(sys.exc_info()[:]))
    def file_upload_u_wash(self):
        
        filename_raw = filedialog.askopenfilename(parent=self)
        self.filename_u_wash = filename_raw
        try:
            self.file_u_meas = load_workbook(filename_raw)
            messagebox.showinfo("Success!", "You have uploaded your U wash file! " )
        except OSError as err:
            messagebox.showwarning("Error", str(err))
        except:
            messagebox.showerror("Unexpected error:", str(sys.exc_info()[:]))
    def file_upload_th_wash(self):
        
        filename_raw = filedialog.askopenfilename(parent=self)
        self.filename_th_wash = filename_raw
        try:
            self.file_th_wash = load_workbook(filename_raw)
            messagebox.showinfo("Success!", "You have uploaded your Th wash file! " )
        except OSError as err:
            messagebox.showwarning("Error", str(err))
        except:
            messagebox.showerror("Unexpected error:", str(sys.exc_info()[:]))
    def file_upload_u_chemblank(self):
    
        filename_raw = filedialog.askopenfilename(parent=self)
        self.filename_u_chemblank = filename_raw
        try:
            self.file_u_chemblank = load_workbook(filename_raw)
            messagebox.showinfo("Success!", "You have uploaded your U chem blank file! " )
        except OSError as err:
            messagebox.showwarning("Error", str(err))
        except:
            messagebox.showerror("Unexpected error:", str(sys.exc_info()[:]))
    def file_upload_th_chemblank(self):
    
        filename_raw = filedialog.askopenfilename(parent=self)
        self.filename_th_chemblank = filename_raw
        try:
            self.file_th_chemblank = load_workbook(filename_raw)
            messagebox.showinfo("Success!", "You have uploaded your Th chem blank file! " )
        except OSError as err:
            messagebox.showwarning("Error", str(err))
        except:
            messagebox.showerror("Unexpected error:", str(sys.exc_info()[:]))
    def file_upload_u_chemblankwash(self):
    
        filename_raw = filedialog.askopenfilename(parent=self)
        self.filename_u_chemblankwash = filename_raw
        try:
            self.file_u_chemblankwash = load_workbook(filename_raw)
            messagebox.showinfo("Success!", "You have uploaded your U chem blank wash file! " )
        except OSError as err:
            messagebox.showwarning("Error", str(err))
        except:
            messagebox.showerror("Unexpected error:", str(sys.exc_info()[:]))
    def file_upload_th_chemblankwash(self):
    
        filename_raw = filedialog.askopenfilename(parent=self)
        self.filename_th_chemblankwash = filename_raw
        try:
            self.file_th_chemblankwash = load_workbook(filename_raw)
            messagebox.showinfo("Success!", "You have uploaded your Th chem blank wash file! " )
        except OSError as err:
            messagebox.showwarning("Error", str(err))
        except:
            messagebox.showerror("Unexpected error:", str(sys.exc_info()[:]))
            
    def file_upload_export(self):
    
        filename_raw = filedialog.askopenfilename(parent=self)
        self.filename_export = filename_raw
        try:
            self.file_export = load_workbook(filename_raw)
            messagebox.showinfo("Success!", "You have uploaded your export file! " )
        except OSError as err:
            messagebox.showwarning("Error", str(err))
        except:
            messagebox.showerror("Unexpected error:", str(sys.exc_info()[:]))
             
             
    def click_submit(self, event=None):

        self.spike_input = self.spikeinput.get()
        spike = self.spike_input
        
        #derives spike value based off dictionary entries
        spike_six_three_dictionary = {"DIII-B":1.008398,"DIII-A": 1.008398,"1I":1.010128,"1H":1.010128}
        spike_six_three_err_dictionary = {"DIII-B": 0.00015, "DIII-A": 0.00015, "1I": 0.00015, "1H": 0.00015}
        spike_three_dictionary = {"DIII-B": 0.78938, "DIII-A": 0.78933, "1I": 0.61351, "1H": 0.78997}
        spike_three_err_dictionary = {"DIII-B": 0.00002, "DIII-A": 0.00002, "1I": 0.00002, "1H": 0.00002}
        spike_nine_dictionary = {"DIII-B": 0.21734, "DIII-A": 0.21705, "1I": 0.177187, "1H": 0.22815}
        spike_nine_err_dictionary = {"DIII-B": 0.00001, "DIII-A": 0.00002, "1I": 0.00001, "1H": 0.00001}
        spike_zero_nine_dictionary = {"DIII-B": 0.0000625, "DIII-A": 0.0000625, "1I": 0.0000402, "1H": 0.0000402}
        spike_zero_nine_err_dictionary = {"DIII-B": 0.000003, "DIII-A": 0.000003, "1I": 0.0000011, "1H": 0.0000011}
        spike_nine_two_dictionary = {"DIII-B": 0.00, "DIII-A": 0.00, "1I": 0.00, "1H": 0.00}
        spike_nine_two_err_dictionary = {"DIII-B": 0.00, "DIII-A": 0.00, "1I": 0.00, "1H": 0.00}
        spike_four_three_dictionary = {"DIII-B": 0.003195, "DIII-A": 0.003195, "1I":0.003180, "1H": 0.003180}
        spike_four_three_err_dictionary= {"DIII-B": 0.000003, "DIII-A": 0.000003, "1I": 0.000003, "1H": 0.000003}
        spike_five_three_dictionary = {"DIII-B": 0.10532, "DIII-A": 0.10532, "1I": 0.10521, "1H":0.10521}
        spike_five_three_err_dictionary = {"DIII-B": 0.00003, "DIII-A": 0.00003, "1I": 0.00003, "1H": 0.00003}
        spike_eight_three_dictionary = {"DIII-B": 0.01680, "DIII-A": 0.01680, "1I": 0.01700, "1H":0.01700 }
        spike_eight_three_err_dictionary = {"DIII-B": 0.00001, "DIII-A": 0.00001,"1I": 0.00001, "1H": 0.00001}

        if spike in spike_six_three_dictionary:
            self.spike_six_three = float(spike_six_three_dictionary[spike]) #spike ratio
        else: 
            messagebox.showwarning("Error!", "No valid spike info entered! ")
        
        if spike in spike_six_three_err_dictionary: 
            self.spike_six_three_err = float(spike_six_three_err_dictionary[spike]) #error of spike ratio
            
        if spike in spike_three_dictionary:
            self.spike_three = float(spike_three_dictionary[spike]) #in pmol/g
        else:pass
    
        if spike in spike_three_err_dictionary:
            self.spike_three_err = float(spike_three_err_dictionary[spike]) #in pmol/g
        else:pass
    
        if spike in spike_nine_dictionary:
            self.spike_nine = float(spike_nine_dictionary[spike]) #in pmol/g
        else: pass
    
        if spike in spike_nine_err_dictionary: 
            self.spike_nine_err = float(spike_nine_err_dictionary[spike]) #in pmol/g
        else: pass
    
        if spike in spike_zero_nine_dictionary:
            self.spike_zero_nine = float(spike_zero_nine_dictionary[spike]) #spike ratio
        else: pass
    
        if spike in spike_zero_nine_err_dictionary:
            self.spike_zero_nine_err = float(spike_zero_nine_err_dictionary[spike]) #error of spike ratio
        else: pass
    
        if spike in spike_nine_two_dictionary: 
            self.spike_nine_two = float(spike_nine_two_dictionary[spike]) #spike ratio
        else: pass
    
        if spike in spike_nine_two_err_dictionary:
            self.spike_nine_two_err = float(spike_nine_two_err_dictionary[spike]) #error of spike ratio
        else: pass
    
        if spike in spike_four_three_dictionary:
            self.spike_four_three = float(spike_four_three_dictionary[spike]) #spike ratio
        else: pass
    
        if spike in spike_four_three_err_dictionary:
            self.spike_four_three_err = float(spike_four_three_err_dictionary[spike]) #error of spike ratio
        else: pass
            
        if spike in spike_five_three_dictionary:
            self.spike_five_three = float(spike_five_three_dictionary[spike]) #spike ratio
        else: pass
        
        if spike in spike_five_three_err_dictionary:
            self.spike_five_three_err = float(spike_five_three_err_dictionary[spike]) #error of spike ratio
        else: pass
        
        if spike in spike_eight_three_dictionary:
            self.spike_eight_three = float(spike_eight_three_dictionary[spike]) #spike ratio
        else: pass
        
        if spike in spike_eight_three_err_dictionary:
            self.spike_eight_three_err = float(spike_eight_three_err_dictionary[spike]) #error of spike ratio
        else: pass
        
        #sample information
        self.AS = self.AS1_input.get()

        self.sample_wt = float(self.samplewt.get())
        #self.filename_export_chem = self.chemFile.get()
        self.chemspike_wt = float(self.chemspikewt.get())
        self.chem_spike_wt = float(self.chemspikewt.get())
        #year run
        self.spike_wt = float(self.spikewt.get())
        self.sample_name = self.samplename.get()
        self.year = float(self.year_input.get())
        #self.chemblank_date = self.chemBlankDate.get()
        self.row = self.row_input.get()
        messagebox.showinfo("Success! " , "You have submitted successfully! ")
            
    def click_cancel(self, event=None):
        messagebox.showwarning("You have clicked cancel", " bye bye! ")
        self.master.destroy()
        
    
    def Age_Calculation(self):
        """
        Input variables for Age Calculation
        """
        
           
        #constants needed in calculations
        wt_229 = 229.031756
        wt_230 = 230.033128
        wt_232 = 232.038051
        wt_233 = 233.039629
        wt_234 = 234.040947
        wt_235 = 235.043924
        wt_236 = 236.045563
        wt_238 = 238.050785
        five_counttime = 0.131
        four_counttime = 1.049
        three_counttime = 0.393
        two_nine_counttime = 1.049
        eight_five_rat = 137.82 #why not 137.83? 
        eight_filament_blank = 0.0001
        eight_filament_blank_err = 0.1
        sample_wt_err = 0.000005
        spike_wt_err = 0.000005
        two_nine_spike = 0.00065
        two_nine_spike_err = 0.00005
        AS_1amu = 1.00E-10
        AS_1amu_err = 0.25 * AS_1amu
        AS_2amu = AS_1amu/2.5
        AS_2amu_err = 0.25 * AS_2amu
        lambda_238 = 0.000000000155125
        lambda_234 = 0.0000028263*0.9985
        lambda_230 = 0.0000091577*1.0014
        threefive_four = 1E-11
        fourfour_four = 1E-11
                              
        """
        Input functions for U, Th, wash, and chem blank values for use in Age Calculation
        """
        
        self.wb_U = Ucalculation(self.spike_input, self.AS, self.filename_u_meas)
        
        self.lstU_Th = self.wb_U.U_normalization_forTh() #provides a list for use in Th normalization
        
        self.lstU_Age = self.wb_U.U_normalization_forAge() #provides a list for use in Age Calculation
        """
            lstU_Age output is a list of the following values: 
                [0]: 235/233 normalized ratio
                [1]: 235/233 normalized ratio error
                [2]: 235/234 normalized and corrected ratio
                [3]: 235/234 normalized and corrected ratio error
                [4]: Unfiltered 233 counts
                [5]: Filtered 234/235 counts
                [6]: Unfiltered 233 mean
        """
        
        self.wb_Th = Thcalculation(self.spike_input, self.AS, self.filename_th_meas, self.lstU_Th)
        
        self.lstTh_Age = self.wb_Th.Th_normalization_forAge() #provides a list to use for Age Calculation
        """
            lstTh_Age provides a list of the following outputs for the Age Calculation: 
                [0]: 230/229 corrected and normalized ratio
                [1]: 230/229 corrected and normalized ratio error
                [2]: 232/229 corrected and normalized ratio
                [3]: 232/229 corrected and normalized ratio error
                [4]: Unfiltered 229 mean
                [5]: Unfiltered 229 counts
        """
        
        self.wb_wash = background_values(self.filename_u_wash, self.filename_th_wash)
        
        self.lstU_wash = self.wb_wash.U_wash() #provides a list of 233, 234, 235 wash values for use in Age Calculation
        """
            lstU_wash provides a list of the following outputs for the Age Calculation: 
                [0]: 233 unfiltered wash in cps
                [1]: 234 unfiltered wash in cps
                [2]: 235 unfiltered wash in cps
                
        """
        
        self.Th_wash = self.wb_wash.Th_wash() #provides the 230 darknoise cpm for use in Age Calculation
        
        self.wb_chemblank = chemblank_values("1H", self.chem_spike_wt,
                                                self.filename_u_chemblankwash, self.filename_th_chemblankwash, 
                                                self.filename_u_chemblank, self.filename_th_chemblank)
        self.lst_chemblank = self.wb_chemblank.blank_calculate() #calculates chem blanks for use in Age Calculation
        """
            lst_chemblank provides a list of the following outputs for the Age Calculation: 
                [0]: 238 chemblank value in pmol
                [1]: 238 chemblank error in pmol
                [2]: 232 chemblank value in pmol
                [3]: 232 chemblank error in pmol
                [4]: 230 chemblank value in fmol
                [5]: 230 chemblank error in fmol
        """
        
        """
        Age Calculation equations
        """
        
        #238 ppb
        
        five_three_max_err = ( (self.lstU_Age[6] * self.lstU_Age[0]) - self.lstU_wash[2] ) / (self.lstU_Age[6] - self.lstU_wash[0])
        
        eight_nmol = (((five_three_max_err -  self.spike_five_three) * self.spike_wt * self.spike_three * eight_five_rat)/1000) /self.sample_wt  
        
        chemblank_corr_238 = ((eight_nmol * self.sample_wt) - (self.lst_chemblank[0]/1000)) / self.sample_wt
        
        filament_blank_corr_238 = chemblank_corr_238 * (1 - (eight_filament_blank/ (self.lstU_Age[6] * five_three_max_err
                                                        * eight_five_rat)))
        
        eight_ppb = filament_blank_corr_238 * wt_238
        
        #238 ppb error
        
        rel_err_1 = (self.lstU_Age[1]/self.lstU_Age[0]) 
        
        three_counting_err = 2 / (self.lstU_Age[6] * self.lstU_Age[4] * three_counttime)**0.5
        
        five_counting_err = 2 / (self.lstU_Age[6] * self.lstU_Age[0] * five_counttime * self.lstU_Age[4])**0.5
        
        rel_err_2 = np.sqrt( (five_counting_err**2) + (three_counting_err**2) + (three_counting_err**2)*(8.0/9.0) )
        
        rel_err_five_three = max(rel_err_1, rel_err_2)
        
        abs_err_five_three = rel_err_five_three * five_three_max_err
        
        eight_nmol_err = eight_nmol * np.sqrt( ((np.sqrt((abs_err_five_three**2) + (0.0000527**2)))/(five_three_max_err - self.spike_five_three))**2 +
                                              (spike_wt_err/self.spike_wt)**2 + 
                                              (self.spike_three_err/self.spike_three)**2 + 
                                              (sample_wt_err/self.sample_wt)**2 )
        
        eight_nmol_err_rel = eight_nmol_err/eight_nmol
        
        chemblank_corr_238_err = np.sqrt( (eight_nmol_err**2) + ((self.lst_chemblank[1]/1000)**2) )
        
        chemblank_corr_238_err_rel = chemblank_corr_238_err / chemblank_corr_238
        
        filament_blank_corr_238_err_rel = np.sqrt( (chemblank_corr_238_err_rel**2) + 
                                                   ( ((eight_filament_blank/(self.lstU_Age[6]*self.lstU_Age[0]*eight_five_rat)) *
                                                     np.sqrt((eight_filament_blank_err/eight_filament_blank)**2 + 
                                                              ((self.lstU_Age[6]*0.05)/self.lstU_Age[6])**2 + 
                                                              ((self.lstU_Age[1]/self.lstU_Age[0])**2)))
                                                    / (1 - (eight_filament_blank/(self.lstU_Age[6]*self.lstU_Age[0]*eight_five_rat))))**2)
        
        filament_blank_corr_238_err = filament_blank_corr_238_err_rel * filament_blank_corr_238
        
        eight_ppb_err = filament_blank_corr_238_err * wt_238
        
        #232 ppt
        
        two_nine_max_err = self.lstTh_Age[2]
        
        two_nine_spike_corr = two_nine_max_err - two_nine_spike
        
        two_nine_chemblank_corr = two_nine_spike_corr - ( self.lst_chemblank[2]/(self.spike_wt * self.spike_nine)  )
        
        two_pmol = two_nine_chemblank_corr * self.spike_wt * self.spike_nine/self.sample_wt
                                                   
        two_ppt = two_pmol * wt_232
        
        #232 ppt error
        
        abs_err_two_nine = self.lstTh_Age[3]
        
        two_nine_spike_corr_err = np.sqrt( (abs_err_two_nine**2) + (two_nine_spike_err **2) )
        
        two_nine_chemblank_corr_err = np.sqrt( (self.lst_chemblank[2]/self.spike_wt*self.spike_nine) * 
                                                np.sqrt( (self.lst_chemblank[3]/self.lst_chemblank[2])**2 + 
                                                      (spike_wt_err/self.spike_wt)**2 + 
                                                      (self.spike_nine_err/self.spike_nine)**2)**2 +
                                                two_nine_spike_corr_err**2)
        
        two_pmol_err = two_pmol * np.sqrt( (two_nine_chemblank_corr_err/two_nine_chemblank_corr)**2 + 
                                           (spike_wt_err/self.spike_wt)**2 + 
                                           (self.spike_nine_err/self.spike_nine)**2 + 
                                           (sample_wt_err/self.sample_wt)**2 )
        
        two_pmol_err_rel = two_pmol_err / two_pmol
        
        two_ppt_err = two_ppt * two_pmol_err_rel
        
        #230 pmol/g
        
        zero_nine_max_err = self.lstTh_Age[0]
        
        zero_nine_spike_corr = zero_nine_max_err - self.spike_zero_nine
        
        zero_nine_AS_corr = zero_nine_spike_corr - AS_1amu - (AS_2amu * self.lstTh_Age[2])
        
        zero_nine_darknoise_corr = zero_nine_AS_corr * (1 - ((self.Th_wash/60)/(self.lstTh_Age[4]*zero_nine_AS_corr)) )
        
        zero_nine_chemblank_corr = zero_nine_darknoise_corr - ( self.lst_chemblank[4]/(self.spike_wt * self.spike_nine * 1000) )
        
        zero_pmol = (zero_nine_chemblank_corr * self.spike_wt * self.spike_nine) / self.sample_wt
        
        #230 pmol/g error
        
        zero_nine_counting_err = self.lstTh_Age[0] * 2 * np.sqrt( (1 / ((self.lstTh_Age[4]/self.lstTh_Age[0])*self.lstTh_Age[5]*two_nine_counttime)) + 
                                                             (1 / (self.lstTh_Age[4]*self.lstTh_Age[5]*two_nine_counttime)  ) )
        
        abs_err_zero_nine = max((zero_nine_max_err*0.00001), zero_nine_counting_err, self.lstTh_Age[1]  )
        
        zero_nine_spike_corr_err = np.sqrt( (abs_err_zero_nine**2) + (0.000003**2) )
        
        zero_nine_AS_corr_err = np.sqrt( (zero_nine_spike_corr_err**2) + (AS_1amu_err**2) + 
                                        ( AS_2amu * self.lstTh_Age[2] * np.sqrt( (AS_2amu_err/AS_2amu)**2 + 
                                         (self.lstTh_Age[3]/self.lstTh_Age[2])**2 ) )**2 )
        
        zero_nine_darknoise_corr_err = zero_nine_darknoise_corr * np.sqrt( (zero_nine_AS_corr_err/zero_nine_AS_corr)**2 + 
                                                                      (((self.Th_wash/60)/(self.lstTh_Age[4]*zero_nine_AS_corr)) * 
                                                                      np.sqrt((0.2**2) + (10/self.lstTh_Age[4])**2 + (zero_nine_AS_corr_err/zero_nine_AS_corr)**2 
                                                                              / (1 - ((self.Th_wash/60)/self.lstTh_Age[4]*zero_nine_AS_corr) ))
                                                                              )**2)
        
        zero_nine_chemblank_corr_err = np.sqrt( zero_nine_darknoise_corr_err**2 +
                                               ( (self.lst_chemblank[4]/(self.spike_wt * self.spike_nine * 1000)) * 
                                                np.sqrt( (self.lst_chemblank[5]/self.lst_chemblank[4])**2 + 
                                                         (spike_wt_err/ self.spike_wt)**2 + 
                                                         (self.spike_nine_err/self.spike_nine)**2 ))**2)
        
        zero_pmol_err = zero_pmol * np.sqrt((zero_nine_chemblank_corr_err/zero_nine_chemblank_corr)**2 + 
                                            (spike_wt_err/self.spike_wt)**2 + 
                                            (self.spike_nine_err/self.spike_nine)**2 + 
                                            (sample_wt_err/self.sample_wt)**2)
        
        zero_pmol_err_rel = zero_pmol_err / zero_pmol
        
        #230/232 atomic ratio
            
        zero_two_atomic = zero_pmol / two_pmol
        
        zero_two_atomic_final = zero_two_atomic * 10**6
        
        #230/232 atomic ratio error 
            
        zero_two_atomic_err_rel = np.sqrt( two_pmol_err_rel**2 + zero_pmol_err_rel**2 )
        
        zero_two_atomic_err = zero_two_atomic_err_rel * zero_two_atomic
        
        zero_two_atomic_err_final = zero_two_atomic_err * 10**6
        
        #d234U measured
            
        zero_nine_measuredU = self.lstU_Age[2] * (1 - self.lstU_wash[1]/(self.lstU_Age[6] * self.lstU_Age[2] * self.lstU_Age[0]))
        
        four_five_wt_avg = zero_nine_measuredU
        
        four_three_max_err = four_five_wt_avg * self.lstU_Age[0]
        
        four_five_tail_corr = four_five_wt_avg * (1 - ((4.0/9.0 * threefive_four) + (5.0/9.0 * fourfour_four)))
        
        four_five_spike_corr_234 = four_five_tail_corr * (1 - (self.spike_four_three/four_three_max_err))
        
        four_five_spike_corr_235 = four_five_spike_corr_234 * (1 / (1- (self.spike_five_three/five_three_max_err)))
        
        four_eight_ppm = (four_five_spike_corr_235 * 10**6) / eight_five_rat
        
        d234U_m = (( four_eight_ppm / ((lambda_238/lambda_234) * 10**6)) - 1) * 1000
        
        #d234U measured error
        
        zero_nine_measuredU_err_rel = self.lstU_Age[3] / zero_nine_measuredU
        
        rel_err_1 = np.sqrt(zero_nine_measuredU_err_rel**2 + (self.lstU_Age[1]/self.lstU_Age[0])**2)
        
        four_counting_err = 2 / (self.lstU_Age[6] * four_three_max_err * four_counttime * self.lstU_Age[4])**0.5
        
        rel_err_2 = np.sqrt(four_counting_err**2 + 2*three_counting_err**2 + (2.0/9.0)*three_counting_err**2)
        
        rel_err_four_three = max(rel_err_1, rel_err_2)
        
        four_five_wt_avg_err_rel = max(zero_nine_measuredU_err_rel**2, 
                                       np.sqrt(four_counting_err**2 + five_counting_err**2 + (2.0/9.0 * three_counting_err**2) ))
        
        four_five_tail_corr_err_rel = np.sqrt((four_five_wt_avg_err_rel**2) + 
                                              (np.sqrt((4.0/9.0 * threefive_four)**2 + (5.0/9.0 * fourfour_four)**2)/
                                               (1 - (4.0/9.0 * threefive_four + 5/9 * fourfour_four)) )**2)
        
        four_five_spike_corr_234_err_rel = np.sqrt((four_five_tail_corr_err_rel**2) + 
                                                   ((self.spike_four_three/four_three_max_err) * np.sqrt(0.002**2 + rel_err_four_three**2) /
                                                    (1 - self.spike_four_three/four_three_max_err))**2)
        
        four_five_spike_corr_235_err_rel = np.sqrt((four_five_spike_corr_234_err_rel**2) + 
                                                   ((self.spike_five_three/five_three_max_err) * np.sqrt(0.0005**2 + (rel_err_five_three/1000)**2) /
                                                    (1 - self.spike_five_three/five_three_max_err))**2)
        
        four_five_spike_corr_235_err = four_five_spike_corr_235_err_rel * four_five_spike_corr_235
        
        four_eight_ppm_err = (four_five_spike_corr_235_err * 10**6)/ eight_five_rat
        
        d234U_m_err = (four_eight_ppm_err / ((lambda_238/lambda_234) * 10**6)) * 1000
        
        #230Th/238U activity ratio
        
        zero_eight_atomic = (zero_pmol/(eight_ppb/wt_238))/1000
        
        zero_eight_activity = zero_eight_atomic * (lambda_230/lambda_238)
        
        #230Th/238U activity ratio error 
        
        zero_eight_atomic_err_rel = np.sqrt(zero_pmol_err_rel**2 + eight_nmol_err_rel **2 )
        
        zero_eight_activity_err = zero_eight_atomic_err_rel * zero_eight_activity
        
        #Uncorrected age calculation and error
        
        age_func = lambda t : zero_eight_activity - (1 - np.exp(-lambda_230*t) + (d234U_m/1000) * 
                                                 (lambda_230/(lambda_230-lambda_234)) * 
                                                 (1 - np.exp((lambda_234 - lambda_230)*t)))
        
        t_initial_guess = 0
        uncorrected_t = fsolve(age_func, t_initial_guess) #returns the value for t at which the solution is 0. This is true of all fsolve functions following this. 
        
        age_func_ThUmax = lambda t : (zero_eight_activity+zero_eight_activity_err) - (1 - np.exp(-lambda_230*t) + (d234U_m/1000) * 
                                                 (lambda_230/(lambda_230-lambda_234)) * 
                                                 (1 - np.exp((lambda_234 - lambda_230)*t)))
        
        uncorrected_ThUmax = fsolve(age_func_ThUmax, t_initial_guess)
        
        age_func_ThUmin = lambda t : (zero_eight_activity-zero_eight_activity_err) - (1 - np.exp(-lambda_230*t) + (d234U_m/1000) * 
                                                 (lambda_230/(lambda_230-lambda_234)) * 
                                                 (1 - np.exp((lambda_234 - lambda_230)*t)))
        
        uncorrected_ThUmin = fsolve(age_func_ThUmin, t_initial_guess)
        
        age_func_d234Umax = lambda t : zero_eight_activity - (1 - np.exp(-lambda_230*t) + ((d234U_m + d234U_m_err)/1000) * 
                                                 (lambda_230/(lambda_230-lambda_234)) * 
                                                 (1 - np.exp((lambda_234 - lambda_230)*t)))
        
        uncorrected_d234Umax = fsolve(age_func_d234Umax, t_initial_guess)
        
        age_func_d234Umin = lambda t : zero_eight_activity - (1 - np.exp(-lambda_230*t) + ((d234U_m - d234U_m_err)/1000) * 
                                                 (lambda_230/(lambda_230-lambda_234)) * 
                                                 (1 - np.exp((lambda_234 - lambda_230)*t)))
        
        uncorrected_d234Umin = fsolve(age_func_d234Umin, t_initial_guess)
        
        uncorrected_t_maxerr = np.sqrt((uncorrected_ThUmax - uncorrected_t)**2 + (uncorrected_d234Umax - uncorrected_t)**2)
        
        uncorrected_t_minerr = np.sqrt((uncorrected_ThUmin - uncorrected_t)**2 + (uncorrected_d234Umin - uncorrected_t)**2)
        
        uncorrected_t_err = (uncorrected_t_maxerr + uncorrected_t_minerr)/2
        
        #Corrected age calculation and error
        
        zero_two_initial = 0.0000044
        zero_two_initial_err = zero_two_initial/2
        
        age_func_corrected_t = lambda t : (((zero_pmol - zero_two_initial*np.exp(-lambda_230*t)*two_pmol) * lambda_230/(filament_blank_corr_238 * 1000 * lambda_238)) - 
                                  (1 - np.exp(-lambda_230 * t) + (d234U_m/1000 * (lambda_230/(lambda_230-lambda_234)) * 
                                  (1 - np.exp((lambda_234-lambda_230)*t)))))
        
        t_initial_guess = 0
        corrected_t = fsolve(age_func_corrected_t, t_initial_guess)
        
        zero_two_initial_now = zero_two_initial * np.exp(-lambda_230 * corrected_t)
        
        zero_two_initial_now_err = zero_two_initial_now * (zero_two_initial_err / zero_two_initial)
        
        corrected_zero_eight_activity = (zero_pmol - zero_two_initial_now*two_pmol) * lambda_230/(filament_blank_corr_238 * 1000 * lambda_238)
        
        corrected_zero_eight_activity_err = corrected_zero_eight_activity * np.sqrt( 
                                                                            (np.sqrt(((zero_two_initial_now * two_pmol) * np.sqrt((zero_two_initial_now_err/zero_two_initial_now)**2 
                                                                                    + (two_pmol_err/two_pmol))**2)**2 + zero_pmol_err **2) / 
                                                                                    (zero_pmol - zero_two_initial_now*two_pmol))**2 +
                                                                                    (filament_blank_corr_238_err/filament_blank_corr_238)**2)
        
        
        age_func_ThUmax = lambda t : (corrected_zero_eight_activity+corrected_zero_eight_activity_err) - (1 - np.exp(-lambda_230*t) + (d234U_m/1000) * 
                                                 (lambda_230/(lambda_230-lambda_234)) * 
                                                 (1 - np.exp((lambda_234 - lambda_230)*t)))
        
        corrected_ThUmax = fsolve(age_func_ThUmax, t_initial_guess)
        
        age_func_ThUmin = lambda t : (corrected_zero_eight_activity-corrected_zero_eight_activity_err) - (1 - np.exp(-lambda_230*t) + (d234U_m/1000) * 
                                                 (lambda_230/(lambda_230-lambda_234)) * 
                                                 (1 - np.exp((lambda_234 - lambda_230)*t)))
        
        corrected_ThUmin = fsolve(age_func_ThUmin, t_initial_guess)
        
        age_func_d234Umax = lambda t : corrected_zero_eight_activity - (1 - np.exp(-lambda_230*t) + ((d234U_m + d234U_m_err)/1000) * 
                                                 (lambda_230/(lambda_230-lambda_234)) * 
                                                 (1 - np.exp((lambda_234 - lambda_230)*t)))
        
        corrected_d234Umax = fsolve(age_func_d234Umax, t_initial_guess)
        
        age_func_d234Umin = lambda t : corrected_zero_eight_activity - (1 - np.exp(-lambda_230*t) + ((d234U_m - d234U_m_err)/1000) * 
                                                 (lambda_230/(lambda_230-lambda_234)) * 
                                                 (1 - np.exp((lambda_234 - lambda_230)*t)))
        
        corrected_d234Umin = fsolve(age_func_d234Umin, t_initial_guess)
        
        age_func_low = lambda t: ((zero_pmol - ((zero_two_initial_now + zero_two_initial_now_err) * np.exp(-lambda_230 * t)) *two_pmol) 
                                * lambda_230/(filament_blank_corr_238 * 1000 * lambda_238)) - (1 - np.exp(-lambda_230*t) + ((d234U_m)/1000) * 
                                             (lambda_230/(lambda_230-lambda_234)) * 
                                             (1 - np.exp((lambda_234 - lambda_230)*t)))
        
        age_func_high = lambda t: ((zero_pmol - ((zero_two_initial_now - zero_two_initial_now_err) * np.exp(-lambda_230 * t)) *two_pmol) 
                                * lambda_230/(filament_blank_corr_238 * 1000 * lambda_238)) - (1 - np.exp(-lambda_230*t) + ((d234U_m)/1000) * 
                                             (lambda_230/(lambda_230-lambda_234)) * 
                                             (1 - np.exp((lambda_234 - lambda_230)*t)))
        
        corrected_age_low = fsolve(age_func_low, t_initial_guess)
        
        corrected_age_high = fsolve(age_func_high, t_initial_guess)
        
        corrected_t_maxerr = np.sqrt((corrected_ThUmax - corrected_t)**2 + (corrected_d234Umax - corrected_t)**2 + (corrected_age_high - corrected_t)**2 )
        
        corrected_t_minerr = np.sqrt((corrected_ThUmin - corrected_t)**2 + (corrected_d234Umin - corrected_t)**2 + (corrected_age_low - corrected_t)**2 )
        
        corrected_t_err = (corrected_t_maxerr + corrected_t_minerr)/2
        
        #Corrected initial d234U and error
        
        d234U_i = d234U_m * np.exp(lambda_234 * corrected_t)
        
        d234U_i_maxerr = np.sqrt( (d234U_m_err * np.exp(lambda_234 * corrected_t))**2 + 
                                 (d234U_m * np.exp((lambda_234 * (corrected_t + corrected_t_maxerr)) - d234U_i))**2)
        
        d234U_i_minerr = np.sqrt( (d234U_m_err * np.exp(lambda_234 * corrected_t))**2 + 
                                 (d234U_m * np.exp((lambda_234 * (corrected_t - corrected_t_minerr)) - d234U_i))**2)
        
        d234U_i_err = (d234U_i_maxerr + d234U_i_minerr)/2
        
        #Corrected age BP
        
        corrected_t_BP = corrected_t - self.year
        
        corrected_t_BP_err = corrected_t_err
        age_file = load_workbook(self.filename_export)
        
        sheet = age_file.get_sheet_by_name('Sheet1')
        row = str(self.row)
        sheet['B' + row] = self.sample_name
        sheet['C' + row] = "{0:.1f}".format(eight_ppb)
        sheet['D' + row] = "± " + "{0:.1f}".format(eight_ppb_err)
        sheet['E' + row] = "{0:.0f}".format(two_ppt)
        sheet['F' + row] = "± " + "{0:.0f}".format(two_ppt_err)
        sheet['G' + row] = "{0:.1f}".format(zero_two_atomic_final)
        sheet['H' + row] = "± " + "{0:.1f}".format(zero_two_atomic_final)
        sheet['I' + row] = "{0:.1f}".format(d234U_m)
        sheet['J' + row] = "± " + "{0:.1f}".format(d234U_m_err)
        sheet['K' + row] = "{0:.5f}".format(zero_eight_activity)
        sheet['L' + row] = "± " + "{0:.5f}".format(zero_eight_activity_err)
        sheet['M' + row] = "%.0f" % uncorrected_t
        sheet['N' + row] = "± %.0f" % uncorrected_t_err
        sheet['O' + row] = "%.0f" % corrected_t
        sheet['P' + row] = "± %.0f" % corrected_t_err
        sheet['Q' + row] = "%.1f" % d234U_i
        sheet['R' + row] = "± %.1f" % d234U_i_err
        sheet['S' + row] = "%.0f" % corrected_t_BP
        sheet['T' + row] = "± %.0f" % corrected_t_BP_err
        
        age_file.save(self.filename_export)
        
            
        
        messagebox.showinfo( "AGE CALCULATION VALUES ",\
        "238 ppb: " + str(eight_ppb) + " ± " + str(eight_ppb_err)+\
        "\n232 ppt: " + str(two_ppt) + " ± " + str(two_ppt_err)+\
        "\n230/232 atomic (10*6) ratio: " + str(zero_two_atomic_final) + " ± " + str(zero_two_atomic_err_final)+\
        "\nd234U measured: " + str(d234U_m) + " ± " + str(d234U_m_err)+\
        "\n230/238 activity ratio: " + str(zero_eight_activity) + " ± " + str(zero_eight_activity_err)+\
        "\n230Th Age uncorrected: %f" % uncorrected_t + " ± %f" % uncorrected_t_err + " yrs"+\
        "\n230Th Age corrected: %f" % corrected_t + " ± %f" % corrected_t_err + " yrs"+\
        "\nd234U initial corrected: %f" % d234U_i + " ± %f" % d234U_i_err+\
        "\n230Th Age corrected: %f" % corrected_t_BP + " ± %f" %corrected_t_BP_err + " yrs BP"+\
        "\nAge Calculation has finished")
        

        messagebox.showinfo("Success! ", "Age calculation finished! ")
    
        

class isofilter():
    def __init__(self, filename,columnletter,filternumber): # input filename and columnletter as strings
        self.column = str(columnletter)+'{}:'+str(columnletter)+'{}'
        self.filename = str(filename)
        self.filternumber = int(filternumber)
        self.workbook = load_workbook(self.filename)
        self.ws = self.workbook.active
        self.totalCounts = 0
        self.mean = 0 
        self.filteredMean = 0
        self.err = 0
        self.criteria = 0
        self.totalCounts_filt = 0
        self.standdev = 0
    
    def getMean(self):
        """
        Code works row by row through specified Excel column, and calculates total mean
        """
        outlist = []
        for row in self.ws.iter_rows(self.column.format(2, self.ws.max_row - 8)):
            for cell in row:
                value = cell.value or cell.value == 0
                if value:
                    outlist.append(cell.value)
                else:
                    outlist.append(np.nan)
        outarray = np.array(outlist, dtype = np.float) 
        self.mean = np.nanmean(a = outarray)
        return self.mean

    def getStanddev(self):
        """
        Code works row by row through specified Excel column, and calculates standard deviation
        """
        outlist = []
        for row in self.ws.iter_rows(self.column.format(2, self.ws.max_row - 8)):
            for cell in row:
                value = cell.value or cell.value == 0
                if value:
                    outlist.append(value)
                else:
                    outlist.append(np.nan)
        outarray = np.array(outlist, dtype = np.float) 
        self.standdev = np.nanstd(a = outarray, ddof = 1)
        return self.standdev
    
    def getCounts(self):
        """
        Code works row by row through specified Excel Column, and determines total number of values present (i.e. cycles)
        """
        total_counts = 0
        for row in self.ws.iter_rows(self.column.format(2, self.ws.max_row - 8)):
            for cell in row:
                value = cell.value or cell.value == 0
                if value:
                    total_counts +=1

        self.totalCounts = total_counts
        return self.totalCounts
        
    def Filtered_mean(self, mean, standdev, counts):
        """
        Code works row by row through specified Excel column, deletes entries that are outside of specified range, 
        and calculates resulting mean
        """
        self.mean = mean
        self.standdev = standdev
        self.totalCounts = counts
        self.standerr = (self.standdev / (self.totalCounts**0.5))
        self.criteria = self.filternumber * self.standerr
        outlist = []
        outcounts = 0
        for row in self.ws.iter_rows(self.column.format(2, self.ws.max_row - 8)):
            for cell in row:
                value = cell.value or cell.value == 0
                if value:
                    if abs(value - self.mean) > self.criteria:
                        outlist.append(np.nan)
                    else:
                        outlist.append(value)
                        outcounts += 1
                else:
                    outlist.append(np.nan)
        outarray = np.array(outlist, dtype = np.float)
        self.filteredMean = np.nanmean(a = outarray)
        return self.filteredMean
    
    def Filtered_err(self, mean, standdev, counts):
        """
        Code works row by row through specified Excel column, deletes entries that are outside of specified range, 
        and calculates resulting 2s counting stantistics error
        """
        self.mean = mean
        self.standdev = standdev
        self.totalCounts = counts
        self.standerr = (self.standdev / (self.totalCounts**0.5))
        self.criteria = self.filternumber * self.standerr
        outlist = []
        outcounts = 0
        for row in self.ws.iter_rows(self.column.format(2, self.ws.max_row - 8)):
            for cell in row:
                value = cell.value or cell.value == 0
                if value:
                    if abs(value - self.mean) > self.criteria:
                        outlist.append(np.nan)
                    else:
                        outlist.append(value)
                        outcounts += 1
                else:
                    outlist.append(np.nan)
        outarray = np.array(outlist, dtype = np.float)
        outstanddev = np.nanstd(a=outarray, ddof = 1)
        self.err = 2 * (outstanddev / (outcounts ** 0.5))
        return self.err
    
    def Filtered_counts(self, mean, standdev, counts):
        """
        Code works row by row through specified Excel column, deletes entries that are outside of specified range, 
        and determines total number of values remaining (i.e. filtered cycles)
        """
        self.mean = mean
        self.standdev = standdev
        self.totalCounts = counts
        self.standerr = (self.standdev / (self.totalCounts**0.5))
        self.criteria = self.filternumber * self.standerr
        outlist = []
        outcounts = 0
        for row in self.ws.iter_rows(self.column.format(2, self.ws.max_row - 8)):
            for cell in row:
                value = cell.value or cell.value == 0
                if value:
                    if abs(value - self.mean) > self.criteria:
                        outlist.append(np.nan)
                    else:
                        outlist.append(value)
                        outcounts += 1
                else:
                    outlist.append(np.nan)
        self.totalCounts_filt = outcounts
        return self.totalCounts_filt
    
class chem_blank():
    
        def __init__(self,filename, columnletter, int_time):
            self.column = str(columnletter)+'{}:'+str(columnletter)+'{}'
            self.filename = str(filename)
            self.workbook = load_workbook(self.filename)
            self.ws = self.workbook.active
            
            int_time = str(int_time)
            
            int_dictionary = {"229":0.131, "230":1.049, "232":0.262, "233":0.131, "234":1.049,
                              "235": 0.262, "236":0.131, "238": 0.262}
            
            if int_time in int_dictionary:
                self.inttime = int_dictionary[int_time]
            else: print "Int_time not available"
                      
        def calc(self):
            """
            Code calculates the mean, total cycles, and 2s counting statistics error of chem blanks, in order
            to use in the Age Calculation
            """
            outlist = []
            outcounts = 0
            for row in self.ws.iter_rows(self.column.format(2, self.ws.max_row - 8)):
                for cell in row:
                    if cell.value: 
                        value = cell.value
                        outlist.append(value)
                        outcounts += 1
                    elif cell.value == 0:
                        value= 0.00
                        outlist.append(value)
                        outcounts +=1
                    else: outlist.append(np.nan)
            outarray = np.array(outlist, dtype = np.float)
            self.mean = np.nanmean(a = outarray)
            standdev = np.nanstd(a = outarray, ddof = 1)
            self.counts = outcounts
            err_abs =  2 * standdev/((self.counts)**0.5)
            err_rel_option1 = err_abs/self.mean
            err_rel_option2 = 2/((self.mean * self.counts*self.inttime)**0.5)
            self.err_rel = max(err_rel_option1, err_rel_option2)
            
            lst_Chem = [self.mean, self.counts, self.err_rel]
        
            return lst_Chem


    
class Ucalculation():
    """
    Class Ucalculation functions as the U sheet in the age calculation spreadsheet. Ucalculation gives outputs for 
    both the Thcalculation function and the Agecalculation function. 
    
    U_normalized_forTh output is a list of the following values: 
        [0]: 236/233 measured ratio
        [1]: 236/233 measured ratio error
        [2]: 235/233 normalized ratio
        [3]: 235/233 normalized ratio error
        [4]: 236/233 corrected ratio
        [5]: 236/233 corrected ratio error
    
    U_normalized_forAge output is a list of the following values: 
        [0]: 235/233 normalized ratio
        [1]: 235/233 normalized ratio error
        [2]: 235/234 normalized and corrected ratio
        [3]: 235/234 normalized and corrected ratio error
        [4]: Unfiltered 233 counts
        [5]: Filtered 234/235 counts
        [6]: Unfiltered 233 mean
    
    """
    def __init__ (self, spike_input, AS_input,filename_input):
        
        spike = str(spike_input)
    
        spike_six_three_dictionary = {"DIII-B":1.008398,"DIII-A": 1.008398,"1I":1.010128,"1H":1.010128}
        
        
        #derives 236/233 value of spike from preset dictionary
        if spike in spike_six_three_dictionary:
            self.spike = float(spike_six_three_dictionary[spike])
       
        #allows you the ability to print as you go
        
    
        
        #AS is the abundant sensitivity 237/238, measured through the AS method on the ICP-MS    
        self.AS = float(AS_input)
        
        #uses the filename given for your U run
        filename = str(filename_input)
                            
        #236/233 filtered measured mean and 2s error
        working = isofilter(filename,"G", 44)
        a = working.getMean()
        b = working.getStanddev()
        c = working.getCounts()
        self.six_three_mean_meas = working.Filtered_mean(a,b,c)
        self.six_three_err_meas = working.Filtered_err(a,b,c)
        
        #235/233 filtered measured mean and 2s error
        working_b = isofilter(filename, "H", 44)
        a = working_b.getMean()
        b = working_b.getStanddev()
        c = working_b.getCounts()
        self.five_three_mean_meas = working_b.Filtered_mean(a,b,c)
        self.five_three_err_meas = working_b.Filtered_err(a,b,c) 
    
        #234/235 filtered measured mean and 2s error
        working_c = isofilter(filename,"I", 44)
        a = working_c.getMean()
        b = working_c.getStanddev()
        c = working_c.getCounts()
        self.four_five_mean_meas = working_c.Filtered_mean(a,b,c)
        self.four_five_err_meas = working_c.Filtered_err(a,b,c) 
        self.four_five_counts = working_c.Filtered_counts(a,b,c)
        
        #233 unfiltered mean and counts
        working_d = isofilter(filename, "C", 44)
        self.three_mean_meas = working_d.getMean()
        self.three_counts = working_d.getCounts()
        
        #constants to be used throughout the class
        self.wt_235 = 235.043924
        self.wt_233 = 233.039629
        self.wt_236 = 236.045563
        self.wt_234 = 234.040947
        self.eight_five_rat = 137.83
        self.AS_six_eight = self.AS/5
        self.AS_four_eight = self.AS/20
        self.eight_five_rat_err_rel = 0.0003
        
        
    def U_normalization_forTh(self):
        """
        Function outputs the measured 236/233 ratio and error, the 236/233 ratio and error corrected for the 238 tail, 
        and the 235/233 normalized ratio and error using the 235/233 corrected ratio and further correcting
        235/233 for mass fractionation in the ICP-MS. These values are used later in the Th_normalization function.
        """
        
        #corrects 236/233 ratio for 238 tail 
        self.six_three_corr = self.six_three_mean_meas * ( 1 - (self.AS_six_eight * self.five_three_mean_meas 
                                                                * self.eight_five_rat/self.spike) )
        
        #provides the ratio that will be used to correct for mass fractionation
        rat = float(np.log(self.wt_235/self.wt_233)/np.log(self.wt_236/self.wt_233))
        
        #corrects for mass fractionation in the ICP-MS
        self.five_three_norm = self.five_three_mean_meas * (self.spike/self.six_three_corr)**rat
        
        #provides relative error constants to be used in this function
        AS_six_eight_err_rel = 0.3
        five_three_err_rel = self.five_three_err_meas/self.five_three_mean_meas
        six_three_err_rel = self.six_three_err_meas/self.six_three_mean_meas
       
        #calculculates the 236/233 corrected error
        self.six_three_corr_err = self.six_three_corr * np.sqrt( six_three_err_rel**2 + 
                                                                ( (self.AS_six_eight * self.five_three_mean_meas 
                                                                   * self.eight_five_rat)/self.spike  
                                                                   * np.sqrt( AS_six_eight_err_rel**2 + five_three_err_rel ** 2 + self.eight_five_rat_err_rel**2 ) 
                                                                   / (1 - (self.AS_six_eight * self.five_three_mean_meas * self.eight_five_rat)
                                                                   / self.spike) ) ** 2 ) 
        #calculates the 236/233 relative corrected error
        self.six_three_corr_err_rel = self.six_three_corr_err/self.six_three_corr
        
        #calculates the 235/233 normalized error
        self.five_three_norm_err = self.five_three_norm * np.sqrt( five_three_err_rel**2 
                                                                  + (2 * (self.six_three_corr_err_rel/3))**2  ) 
       
        #if you have chosen to print as you go, this will print when the function is finished
        '''
        while self.inquiry:
            print "RESULTS FOR TH FILTERING FROM U:"
            print "236/233 measured ratio: " + str(self.six_three_mean_meas) + " ± " + str(self.six_three_err_meas)
            print "235/233 normalized ratio: " + str(self.five_three_norm) + " ± " + str(self.five_three_norm_err)
            print "236/233 corrected ratio: " + str(self.six_three_corr) + " ± " + str(self.six_three_corr_err)
            break
        '''
        #a list of your outputs is created and returned, to be used in the Th functions
        lstU_Th = [self.six_three_mean_meas, self.six_three_err_meas, self.five_three_norm, 
                 self.five_three_norm_err, self.six_three_corr,self.six_three_corr_err ]
        
        return lstU_Th
    
    def U_normalization_forAge(self):
        """
        Function outputs the 235/233 normalized ratio and error, the 235/234 normalized and corrected ratio and error, 
        the unfiltered number of cycles for 233 and mean value and the filtered number of cycles 234/235. These values will be used
        later in the Age Calculation.
        """
        #calculates constants that will be used to calculate normalized 234/235
        four_five_err_rel = self.four_five_err_meas / self.four_five_mean_meas
        
        rat = float(np.log(self.wt_234/self.wt_235)/np.log(self.wt_236/self.wt_233))
        
        #normalizes the 234/235 ratio by correcting for mass fractionation and calculates the resulting error
        self.four_five_norm = self.four_five_mean_meas * (self.spike/self.six_three_corr)**rat
        
        self.four_five_norm_err = self.four_five_norm * np.sqrt( four_five_err_rel**2 + 
                                                                (self.six_three_corr_err_rel/3)**2 )
        
        #calculates constants that will be used to calculated corrected 234/235
        AS_four_eight_err_rel = 0.3
        four_five_norm_err_rel = self.four_five_norm_err/self.four_five_norm
        
        #corrects the normalized 234/235 ratio for 238 tail and calculated the resulting error
        self.four_five_normcorr = self.four_five_norm * (1 - ( self.eight_five_rat 
                                                              * self.AS_four_eight/ self.four_five_norm ))

        self.four_five_normcorr_err = self.four_five_normcorr * np.sqrt( four_five_norm_err_rel**2 + 
                                                                        ( (self.eight_five_rat * self.AS_four_eight / self.four_five_norm) *
                                                                         np.sqrt( self.eight_five_rat_err_rel**2 + AS_four_eight_err_rel**2 + four_five_norm_err_rel**2 )
                                                                         / (1 - ( self.eight_five_rat * self.AS_four_eight/ self.four_five_norm)) ) **2 ) 
        
        self.three_mean_meas = int(self.three_mean_meas)
        
        #if you have chosen to print as you go, this will print when the function is finished
        '''
        while self.inquiry:
            print "RESULTS FOR AGE CALC FROM U:"
            print "235/233 normalized ratio : " + str(self.five_three_norm) + " ± " + str(self.five_three_norm_err)
            print "234/235 normalized and corrected ratio: " + str(self.four_five_normcorr) + " ± " + str(self.four_five_normcorr_err)
            print "Unfiltered cycles of 233: " + str(self.three_counts)
            print "Filtered cycles of 234/235: " + str(self.four_five_counts)
            print "Unfiltered mean 233U cps : " + str(self.three_mean_meas)
            break
        '''
        #a list of your outputs is created and returned, to be used in the Age functions
        lstU_Age = [self.five_three_norm, self.five_three_norm_err, self.four_five_normcorr, 
                    self.four_five_normcorr_err, self.three_counts, self.four_five_counts, self.three_mean_meas] 
       
        return lstU_Age
    
class Thcalculation():
    """
    Class Thcalculation functions as the Th sheet in the age calculation spreadsheet. Thcalculation gives outputs for 
    the Agecalculation function, and needs to be provided inputs from the Ucalculation class U_normalization_forTh function.
    
    Th_normalization_forAge output is a list of the following values: 
        [0]: 230/229 corrected and normalized ratio
        [1]: 230/229 corrected and normalized ratio error
        [2]: 232/229 corrected and normalized ratio
        [3]: 232/229 corrected and normalized ratio error
        [4]: Unfiltered 229 mean
        [5]: Unfiltered 229 counts
        
    """
    
    def __init__ (self, spike_input, AS_input, filename_input,lstU_Th):
        
        spike = str(spike_input)
    
        spike_six_three_dictionary = {"DIII-B":1.008398,"DIII-A": 1.008398,"1I":1.010128,"1H":1.010128}
        
        
        #derives 236/233 value of spike from preset dictionary
        if spike in spike_six_three_dictionary:
            self.spike = float(spike_six_three_dictionary[spike])
    
            
        #AS is the abundant sensitivity 237/238, measured through the AS method on the ICP-MS    
        self.AS = float(AS_input)
        
        #uses the filename given for your Th run
        filename = str(filename_input)
        
        #Compiles the values of the lstU_Th provided by your U_normalization_forTh function
        self.six_three_mean_meas = lstU_Th[0]
        self.six_three_err_meas = lstU_Th[1]
        self.five_three_norm = lstU_Th[2]
        self.five_three_norm_err = lstU_Th[3]
        self.six_three_corr = lstU_Th[4]
        self.six_three_corr_err = lstU_Th[5]
        
        #Note: Hai's macro only filters 230/229 column
        
        #230/232 filtered measured mean and 2s error
        working = isofilter(filename,"G", 28)
        self.zero_two_mean_meas = working.getMean()/1.02
        self.zero_two_counts = working.getCounts()
        self.zero_two_standdev_meas = working.getStanddev()
        self.zero_two_rel_err_meas = (2 * self.zero_two_standdev_meas/(self.zero_two_counts**0.5))/self.zero_two_mean_meas
        self.zero_two_rel_err = max(self.zero_two_rel_err_meas, 0.02)
        self.zero_two_err_meas = self.zero_two_mean_meas * self.zero_two_rel_err
        
        #230/229 filtered measured mean and 2s error
        working_b = isofilter(filename, "E", 28)
        a = working_b.getMean()
        b = working_b.getStanddev()
        c = working_b.getCounts()
        self.zero_nine_mean_meas = working_b.Filtered_mean(a,b,c)
        self.zero_nine_err_meas = working_b.Filtered_err(a,b,c)
        
        #232/229 filtered measured mean and 2s error
        working_c = isofilter(filename, "F", 28)
        self.nine_two_mean_meas = working_c.getMean()
        self.two_nine_mean_meas = 1 / (self.nine_two_mean_meas/1.02)
        self.two_nine_counts = working.getCounts()
        self.nine_two_standdev_meas = working_c.getStanddev()
        self.nine_two_rel_err_meas = (2 * self.nine_two_standdev_meas/(self.two_nine_counts**0.5))/self.nine_two_mean_meas
        self.two_nine_rel_err = max(self.nine_two_rel_err_meas, 0.02)
        self.two_nine_err_meas = self.two_nine_mean_meas * self.two_nine_rel_err
        
        #229 unfiltered mean and counts
        working_d = isofilter(filename, "C", 28)
        self.nine_mean_meas = working_d.getMean()
        self.nine_counts = working_d.getCounts()
        
        #constants to be used throughout the class
        self.wt_233 = 233.039629
        self.wt_236 = 236.045563
        self.wt_229 = 229.031756
        self.wt_230 = 230.033128
        self.wt_232 = 232.038051
        self.AS_zero_nine = self.AS
        self.AS_zero_two = self.AS_zero_nine / 5
        self.AS_two_nine = self.AS_zero_two / 3
        self.eight_five_rat = 137.83
        self.eight_five_rat_err_rel = 0.0003
        
    def Th_normalization_forAge(self):
        
        #corrects the 230/229 and 232/229 ratios for both the 232 and 229 tails
        self.zero_nine_corr = self.zero_nine_mean_meas * (1 - self.AS_zero_two/self.zero_two_mean_meas) * (1 - self.AS_zero_nine)
        
        self.two_nine_corr = self.two_nine_mean_meas * (1 / (1 - (self.AS_two_nine * self.two_nine_mean_meas)))
        
        #constants needed for error calculations
        self.zero_nine_rel_err = self.zero_nine_err_meas/self.zero_nine_mean_meas
        self.AS_zero_two_rel_err = 0.3
        self.AS_zero_nine_rel_err = 0.3
        self.AS_two_nine_rel_err = 0.3
        
        #errors for corrected 230/229 and 232/229 ratios
        self.zero_nine_corr_err = self.zero_nine_corr * ( self.zero_nine_rel_err**2 + 
                                                         ( (self.AS_zero_two/self.zero_two_mean_meas) *  
                                                                (self.AS_zero_two_rel_err**2 + self.zero_two_rel_err**2)**0.5
                                                                 / (1 - self.AS_zero_two/self.zero_two_mean_meas))**2 
                                                         + ( self.AS_zero_nine * self.AS_zero_nine_rel_err/(1 - self.AS_zero_nine) )**2) ** 0.5
       
        self.two_nine_corr_err = self.two_nine_corr * ( self.two_nine_rel_err**2 + 
                                                      ( ((self.AS_two_nine_rel_err**2 + self.two_nine_rel_err**2)**0.5)  
                                                       * (self.AS_two_nine * self.two_nine_mean_meas)/
                                                         (1 - (self.AS_two_nine * self.two_nine_mean_meas))**2) ** 2
                                                       ) ** 0.5
        
        #constant needed for normalization
        
        rat_1 = np.log(self.wt_230/self.wt_229) / np.log(self.wt_236/self.wt_233)
        
        rat_2 = np.log(self.wt_232/self.wt_229) / np.log(self.wt_236/self.wt_233)
        

        #normalizes for corrected 230/229 and 232/229 ratios for mass fractionation
        
        self.zero_nine_corrnorm = self.zero_nine_corr * ((self.spike / self.six_three_corr)**rat_1)
        
        self.two_nine_corrnorm = self.two_nine_corr * ((self.spike / self.six_three_corr)**rat_2)
        
        #constants needed for error calculations
        self.zero_nine_corr_rel_err = self.zero_nine_corr_err / self.zero_nine_corr
        self.six_three_corr_rel_err = self.six_three_corr_err / self.six_three_corr
        self.two_nine_corr_rel_err = self.two_nine_corr_err / self.two_nine_corr
        
        #errors for normalized 230/229 and 232/229 ratios 
        self.zero_nine_corrnorm_err = self.zero_nine_corrnorm * ( self.zero_nine_corr_rel_err**2
                                                                 + ( (self.six_three_corr_rel_err/3)**2 )
                                                                 )**0.5
        
        self.two_nine_corrnorm_err = self.two_nine_corrnorm * ( self.two_nine_corr_rel_err**2
                                                               +  self.six_three_corr_rel_err**2  
                                                               )**0.5
        
        self.nine_mean_meas = int(self.nine_mean_meas)
        
        #if you have chosen to print as you go, this will print when the function is finished
        '''
        while self.inquiry:
            print "RESULTS FOR AGE CALC FROM TH:"
            print "230/229 corrected and normalized ratio : " + str(self.zero_nine_corrnorm)  + " ± " + str(self.zero_nine_corrnorm_err) 
            print "232/229 corrected and normalized ratio: " + str(self.two_nine_corrnorm) + " ± " + str(self.two_nine_corrnorm_err)
            print "Unfiltered mean 229Th cps: " + str(self.nine_mean_meas)    
            print "Unfiltered cycles of 229: " + str(self.nine_counts)
            break
        '''
        #a list of your outputs is created and returned, to be used in the Age functions
        lstTh_age = [self.zero_nine_corrnorm, self.zero_nine_corrnorm_err, self.two_nine_corrnorm,
                     self.two_nine_corrnorm_err, self.nine_mean_meas, self.nine_counts]
        
        return lstTh_age
    
    
class background_values():
    
    def __init__(self, U_file, Th_file):
        
        #inquiry = str(inquiry_input)
        
        #allows you the ability to print as you go
        '''
        if inquiry.lower() == "y":
            self.inquiry = True
        else:
            self.inquiry = False
            print "Program background_values ran without printing"
        '''
        #uses the filename given for your U wash
        self.filename_U = str(U_file)
        
        #uses the filename give for your Th wash
        self.filename_Th = str(Th_file)
        
    def U_wash(self):
        """
        U_wash provides a list the following outputs for the Age Calculation: 
            [0]: 233 unfiltered wash in cps
            [1]: 234 unfiltered wash in cps
            [2]: 235 unfiltered wash in cps
            
        """
        #233 wash value
        working_a = isofilter(self.filename_U,"C", 44)
        self.three_wash = working_a.getMean()
        
        #234 wash value
        working_b = isofilter(self.filename_U,"D", 44)
        self.four_wash = working_b.getMean()
        
        #235 wash value
        working_c = isofilter(self.filename_U,"E", 44)
        self.five_wash = working_c.getMean()
        '''
        while self.inquiry: 
            print "BACKGROUND VALUES FOR AGE CALC:"
            print "233 wash value: " + str(self.three_wash) + ' cps'
            print "234 wash value: " + str(self.four_wash) + ' cps'
            print "235 wash value: " + str(self.five_wash) + ' cps'
            break
        '''
        lstU_wash = [self.three_wash, self.four_wash, self.five_wash]
        
        return lstU_wash
    
    def Th_wash(self):
        """
        Th_wash provides the following outputs for the Age Calculation: 
            230 unfiltered wash in cpm 
                
        """
        #230 wash value
        working_a = isofilter(self.filename_Th, "D", 28)
        self.zero_wash = working_a.getMean()
        
        #calculate "darknoise" value for Age calculation
        self.darknoise = self.zero_wash * 60
        '''
        while self.inquiry:
            print "230 wash value/darknoise: " + str(self.darknoise) + ' cpm'
            break
        '''   
        return self.darknoise

class chemblank_values():
    """
    Calculates chem blank values for use in Age Calculation spreadsheet
    
    lstChemblank provides a list of the following outputs: 
        [0]: 238 chemblank value in pmol
        [1]: 238 chemblank error in pmol
        [2]: 232 chemblank value in pmol
        [3]: 232 chemblank error in pmol
        [4]: 230 chemblank value in fmol
        [5]: 230 chemblank error in fmol
        
    Provides the option of printing excel file with chem blank information for personal use.
    """
    
    def __init__(self, spike_input, chem_spike_wt, U_wash, Th_wash, U_chemblank, Th_chemblank):
        
        spike = str(spike_input)
        
        #derives spike value based off dictionary entries
        spike_six_three_dictionary = {"DIII-B":1.008398,"DIII-A": 1.008398,"1I":1.010128,"1H":1.010128}
        spike_six_three_err_dictionary = {"DIII-B": 0.00015, "DIII-A": 0.00015, "1I": 0.00015, "1H": 0.00015}
        spike_three_dictionary = {"DIII-B": 0.78938, "DIII-A": 0.78933, "1I": 0.61351, "1H": 0.78997}
        spike_nine_dictionary = {"DIII-B": 0.21734, "DIII-A": 0.21705, "1I": 0.177187, "1H": 0.22815}
        spike_zero_nine_dictionary = {"DIII-B": 0.0000625, "DIII-A": 0.0000625, "1I": 0.0000402, "1H": 0.0000402}
        spike_zero_nine_err_dictionary = {"DIII-B": 0.000003, "DIII-A": 0.000003, "1I": 0.0000011, "1H": 0.0000011}
        spike_nine_two_dictionary = {"DIII-B": 0.00, "DIII-A": 0.00, "1I": 0.00, "1H": 0.00}
        spike_nine_two_err_dictionary = {"DIII-B": 0.00, "DIII-A": 0.00, "1I": 0.00, "1H": 0.00}
        spike_four_three_dictionary = {"DIII-B": 0.003195, "DIII-A": 0.003195, "1I":0.003180, "1H": 0.003180}
        spike_four_three_err_dictionary= {"DIII-B": 0.000003, "DIII-A": 0.000003, "1I": 0.000003, "1H": 0.000003}
        spike_five_three_dictionary = {"DIII-B": 0.10532, "DIII-A": 0.10532, "1I": 0.10521, "1H":0.10521}
        spike_five_three_err_dictionary = {"DIII-B": 0.00003, "DIII-A": 0.00003, "1I": 0.00003, "1H": 0.00003}
        spike_eight_three_dictionary = {"DIII-B": 0.01680, "DIII-A": 0.01680, "1I": 0.01700, "1H":0.01700 }
        spike_eight_three_err_dictionary = {"DIII-B": 0.00001, "DIII-A": 0.00001,"1I": 0.00001, "1H": 0.00001}
        
        if spike in spike_six_three_dictionary:
            self.spike_six_three = float(spike_six_three_dictionary[spike]) #spike ratio
        else: 
            print 'ERROR: You did not enter a valid spike option'
        
        if spike in spike_six_three_err_dictionary: 
            self.spike_six_three_err = float(spike_six_three_err_dictionary[spike]) #error of spike ratio
            
        if spike in spike_three_dictionary:
            self.spike_three = float(spike_three_dictionary[spike]) #in pmol/g
        else:pass

        if spike in spike_nine_dictionary:
            self.spike_nine = float(spike_nine_dictionary[spike]) #in pmol/g
        else: pass
    
        if spike in spike_zero_nine_dictionary:
            self.spike_zero_nine = float(spike_zero_nine_dictionary[spike]) #spike ratio
        else: pass
    
        if spike in spike_zero_nine_err_dictionary:
            self.spike_zero_nine_err = float(spike_zero_nine_err_dictionary[spike]) #error of spike ratio
        else: pass
    
        if spike in spike_nine_two_dictionary: 
            self.spike_nine_two = float(spike_nine_two_dictionary[spike]) #spike ratio
        else: pass
    
        if spike in spike_nine_two_err_dictionary:
            self.spike_nine_two_err = float(spike_nine_two_err_dictionary[spike]) #error of spike ratio
        else: pass
    
        if spike in spike_four_three_dictionary:
            self.spike_four_three = float(spike_four_three_dictionary[spike]) #spike ratio
        else: pass
    
        if spike in spike_four_three_err_dictionary:
            self.spike_four_three_err = float(spike_four_three_err_dictionary[spike]) #error of spike ratio
        else: pass
        
        if spike in spike_five_three_dictionary:
            self.spike_five_three = float(spike_five_three_dictionary[spike]) #spike ratio
        else: pass
    
        if spike in spike_five_three_err_dictionary:
            self.spike_five_three_err = float(spike_five_three_err_dictionary[spike]) #error of spike ratio
        else: pass
    
        if spike in spike_eight_three_dictionary:
            self.spike_eight_three = float(spike_eight_three_dictionary[spike]) #spike ratio
      
    
        if spike in spike_eight_three_err_dictionary:
            self.spike_eight_three_err = float(spike_eight_three_err_dictionary[spike]) #error of spike ratio
        else: pass
        
        '''
        inquiry = str(inquiry_input)
         
        #allows you to print as you go
        if inquiry.lower() == "y":
            self.inquiry = True
        else:
            self.inquiry = False
            print "Program chemblank_values ran without printing"
        '''
        #spike weight used for chem blank
        self.spike_wt = float(chem_spike_wt)
        
        #spike in chem blank (pmol)
        self.spike_three_used = self.spike_three * self.spike_wt
        self.spike_nine_used = self.spike_nine * self.spike_wt
        
        #allows you to write new chem blank file
        self.chemblank_file = True
       # self.chemblank_filename = str(chemblank_filename)
            
        #file inputs
        self.U_wash = U_wash
        self.Th_wash = Th_wash
        self.U_file = U_chemblank
        self.Th_file = Th_chemblank
        
        #constants to be used throughout the class
        self.wt_229 = 229.031756
        self.wt_230 = 230.033128
        self.wt_232 = 232.038051
        self.wt_233 = 233.039629
        self.wt_234 = 234.040947
        self.wt_235 = 235.043924
        self.wt_236 = 236.045563
        self.wt_238 = 238.050785
    
    def blank_calculate(self):
        """
        Calculates wash and chem blank values for all isotopes. Returns a list of three values:
            1. mean
            2. counts
            3. relative error
        Input needed: file name, column letter, isotope analyzed
        """
        
        """
        Th wash and chem blank values
        """
        #wash 229 Th
        working_a = chem_blank(self.Th_wash, "C", "229")
        nine_wash = working_a.calc()
        
        #chem blank 229 Th
        working_b = chem_blank(self.Th_file, "C", "229")
        nine = working_b.calc()
        
        #wash 230 Th
        working_c = chem_blank(self.Th_wash, "D", "230")
        zero_wash = working_c.calc()
        
        #chem blank 230 Th
        working_d = chem_blank(self.Th_file, "D", "230")
        zero = working_d.calc()
        
        #wash 232 Th
        working_e = chem_blank(self.Th_wash, "E", "232")
        two_wash = working_e.calc()
        
        #chem blank 232Th
        working_f = chem_blank(self.Th_file, "E", "232")
        two = working_f.calc()
        
        """
        U wash and chem blank values
        """
        
        #wash 233U
        working_g = chem_blank(self.U_wash, "D", "233")
        three_wash = working_g.calc()
        
        #chem blank 233U
        working_h = chem_blank(self.U_file, "D", "233")
        three = working_h.calc()
        
        #wash 234U
        working_i = chem_blank(self.U_wash, "E", "234")
        four_wash = working_i.calc()
        
        #chem blank 234U
        working_j = chem_blank(self.U_file, "E", "234")
        four = working_j.calc()
        
        #wash 235U
        working_k = chem_blank(self.U_wash, "F", "235")
        five_wash = working_k.calc()
        
        #chem blank 235U
        working_l = chem_blank(self.U_file, "F", "235")
        five = working_l.calc()
        
        #wash 236U
        working_m = chem_blank(self.U_wash, "G", "236")
        six_wash = working_m.calc()
        
        #chem blank 236U
        working_n = chem_blank(self.U_file, "G", "236")
        six = working_n.calc()
        
        #wash 238U
        working_o = chem_blank(self.U_wash, "H", "238")
        eight_wash = working_o.calc()
        
        #chem blank 238U
        working_p = chem_blank(self.U_file, "H", "238")
        eight = working_p.calc()
        
        """
        Calculates signal isotopic ratio and 2s error
        
        Note: [0]: mean, [1]: counts, [2] = 2s rel error
        
        """
        #230/229
        zero_nine = (zero[0] - zero_wash[0]) / (nine[0] - nine_wash[0])
        zero_nine_err = np.sqrt( ((zero[0]*zero[2])**2/(nine[0] - nine_wash[0])**2) + 
                                 ((zero_wash[0]*zero_wash[2])**2/(nine_wash[0]-nine[0])**2) + 
                                 ((nine[0]*nine[2])**2 * ((zero_wash[0]-zero[0])/((nine[0]-nine_wash[0])**2))**2) + 
                                 ((nine_wash[0]*nine_wash[2])**2 * ((zero[0]-zero_wash[0])/((nine[0]-nine_wash[0])**2))**2)
                                 )/zero_nine
        
        #229/232 
        nine_two = (nine[0] - nine_wash[0]) / (two[0] - two_wash[0])
        nine_two_err = np.sqrt( ((nine[0]*nine[2])**2/(two[0] - two_wash[0])**2) + 
                                 ((nine_wash[0]*nine_wash[2])**2/(two_wash[0]-two[0])**2) + 
                                 ((two[0]*two[2])**2 * ((nine_wash[0]-nine[0])/((two[0]-two_wash[0])**2))**2) + 
                                 ((two_wash[0]*two_wash[2])**2 * ((nine[0]-nine_wash[0])/((two[0]-two_wash[0])**2))**2)
                                 )/nine_two
       
        #234/233
        four_three = (four[0] - four_wash[0])/(three[0] - three_wash[0])
        four_three_err = np.sqrt( ((four[0]*four[2])**2/(three[0] - three_wash[0])**2) + 
                                 ((four_wash[0]*four_wash[2])**2/(three_wash[0]-three[0])**2) + 
                                 ((three[0]*three[2])**2 * ((four_wash[0]-four[0])/((three[0]-three_wash[0])**2))**2) + 
                                 ((three_wash[0]*three_wash[2])**2 * ((four[0]-four_wash[0])/((three[0]-three_wash[0])**2))**2)
                                 )/four_three
        
        #235/233
        five_three = (five[0] - five_wash[0])/(three[0] - three_wash[0])
        five_three_err = np.sqrt( ((five[0]*five[2])**2/(three[0] - three_wash[0])**2) + 
                                 ((five_wash[0]*five_wash[2])**2/(three_wash[0]-three[0])**2) + 
                                 ((three[0]*three[2])**2 * ((five_wash[0]-five[0])/((three[0]-three_wash[0])**2))**2) + 
                                 ((three_wash[0]*three_wash[2])**2 * ((five[0]-five_wash[0])/((three[0]-three_wash[0])**2))**2)
                                 )/five_three
        
        #236/233 
        six_three = (six[0] - six_wash[0])/(three[0] - three_wash[0])
        
        #238/233
        eight_three = (eight[0] - eight_wash[0])/(three[0] - three_wash[0])
        eight_three_err = np.sqrt( ((eight[0]*eight[2])**2/(three[0] - three_wash[0])**2) + 
                                 ((eight_wash[0]*eight_wash[2])**2/(three_wash[0]-three[0])**2) + 
                                 ((three[0]*three[2])**2 * ((eight_wash[0]-eight[0])/((three[0]-three_wash[0])**2))**2) + 
                                 ((three_wash[0]*three_wash[2])**2 * ((eight[0]-eight_wash[0])/((three[0]-three_wash[0])**2))**2)
                                 )/eight_three
        
        """
        Corrects signal isotopic ratios for fractionation
    
        """
        #230/229 fract. corrected
        zero_nine_corr = zero_nine * (self.spike_six_three/six_three)**(np.log(self.wt_230/self.wt_229)/np.log(self.wt_236/self.wt_233))
        
        #229/232 fract. corrected
        nine_two_corr = nine_two * (self.spike_six_three/six_three)**(np.log(self.wt_229/self.wt_232)/np.log(self.wt_236/self.wt_233))
        
        #234/233 fract. corrected
        
        four_three_corr = four_three * (self.spike_six_three/six_three)**(np.log(self.wt_234/self.wt_233)/np.log(self.wt_236/self.wt_233))
        
        #235/233 fract. corrected
        
        five_three_corr = five_three * (self.spike_six_three/six_three)**(np.log(self.wt_235/self.wt_233)/np.log(self.wt_236/self.wt_233))
        
        #238/233 fract. corrected
        
        eight_three_corr = eight_three * (self.spike_six_three/six_three)**(np.log(self.wt_238/self.wt_233)/np.log(self.wt_236/self.wt_233))
        
        
        """
        2s relative spike errors
        """
        
        zero_nine_spike_err = self.spike_zero_nine_err/self.spike_zero_nine
        
        nine_two_spike_err = 0 #may need to change for different spikes
        
        four_three_spike_err = self.spike_four_three_err/self.spike_four_three
        
        five_three_spike_err = self.spike_five_three_err/self.spike_five_three
        
        eight_three_spike_err = self.spike_eight_three_err/self.spike_eight_three
        
        """
        Isotopic ratio and 2s error corrected for fractionation and spike signal
        """
        
        #230/229 fract. corrected and spike corrected
        
        zero_nine_corr_spike = zero_nine_corr - self.spike_zero_nine
        
        zero_nine_corr_spike_err = np.sqrt((zero_nine * zero_nine_err)**2 + (self.spike_zero_nine * zero_nine_spike_err)**2)/ abs(zero_nine_corr_spike)
        
        #229/232 fract. corrected and spike corrected
        
        nine_two_corr_spike = nine_two_corr - self.spike_nine_two
        
        nine_two_corr_spike_err = np.sqrt((nine_two * nine_two_err)**2 + (self.spike_nine_two * nine_two_spike_err)**2)/ abs(nine_two_corr_spike)
        
        #234/233 fract. corrected and spike corrected
        
        four_three_corr_spike = four_three_corr - self.spike_four_three
        
        four_three_corr_spike_err = np.sqrt((four_three * four_three_err)**2 + (self.spike_four_three * four_three_spike_err)**2)/ abs(four_three_corr_spike)
        
        #235/233 fract. corrected and spike corrected
        
        five_three_corr_spike = five_three_corr - self.spike_five_three
        
        five_three_corr_spike_err = np.sqrt((five_three * five_three_err)**2 + (self.spike_five_three * five_three_spike_err)**2)/ abs(five_three_corr_spike)
        
        #238/233 fact. corrected and spike corrected 
        
        eight_three_corr_spike = eight_three_corr - self.spike_eight_three
        
        eight_three_corr_spike_err = np.sqrt((eight_three * eight_three_err)**2 + (self.spike_eight_three * eight_three_spike_err)**2)/ abs(eight_three_corr_spike)
        
        
       
            
            
        self.zero_chemblank = ((self.spike_nine_used * zero_nine_corr_spike)/(10**12))* self.wt_230 * (10**18) #in ag
        self.zero_chemblank_err = abs(self.zero_chemblank * zero_nine_corr_spike_err)
            
        self.two_chemblank = ((self.spike_nine_used / nine_two_corr_spike)/(10**12))* self.wt_232 * (10**15) #in fg
        self.two_chemblank_err = abs(self.two_chemblank * nine_two_corr_spike_err)
            
        self.four_chemblank = ((self.spike_three_used * four_three_corr_spike)/(10**12))* self.wt_234 * (10**18) #in ag
        self.four_chemblank_err = abs(self.four_chemblank * four_three_corr_spike_err)
            
        self.five_chemblank = ((self.spike_three_used * five_three_corr_spike)/(10**12))* self.wt_235 * (10**15) #in fg
        self.five_chemblank_err = abs(self.five_chemblank * five_three_corr_spike_err)
            
        self.eight_chemblank = ((self.spike_three_used * eight_three_corr_spike)/(10**12))* self.wt_238 * (10**15) #in fg
        self.eight_chemblank_err = abs(self.eight_chemblank * eight_three_corr_spike_err)
        '''
        sample_name = sample_name
        chemblank_date = chemblank_date
       
        data = {'1_info': pd.Series([sample_name, chemblank_date], index = ['1_filename', '2_date']),
                '230Th': pd.Series([self.zero_chemblank, self.zero_chemblank_err, 'ag'], index = ['3_chemistry blank', '4_2s abs. err', '5_units']),
                '232Th': pd.Series([self.two_chemblank, self.two_chemblank_err, 'fg'], index = ['3_chemistry blank', '4_2s abs. err', '5_units']),
                '234U': pd.Series([self.four_chemblank, self.four_chemblank_err, 'ag'], index = ['3_chemistry blank', '4_2s abs. err', '5_units']),
                '235U': pd.Series([self.five_chemblank, self.five_chemblank_err, 'fg'], index = ['3_chemistry blank', '4_2s abs. err', '5_units']),
                '238U': pd.Series([self.eight_chemblank, self.eight_chemblank_err, 'fg'], index = ['3_chemistry blank', '4_2s abs. err', '5_units'])}   
        
        df = pd.DataFrame(data)
        
        writer = pd.ExcelWriter(self.chemblank_filename, engine = 'openpyxl')
        
        df.to_excel(writer)
        
        writer.save()
        '''    
        #messagebox.showinfo("Chemblank data file saved ! ", "chemblank data file name: "+ str(self.chemblank_filename))
            
       
        """
        Final calculation for chem blank values for input into Age Calculation
        """
        
        #238 chem blank value and error in pmol
        
        self.chem_blank_eight = self.spike_three_used * eight_three_corr_spike 
        
        self.chem_blank_eight_err = abs(self.chem_blank_eight * eight_three_corr_spike_err) 
        
        #232 chem blank value and error in pmol
        
        self.chem_blank_two = self.spike_nine_used * nine_two_corr_spike 
        
        self.chem_blank_two_err = abs(self.chem_blank_two * nine_two_corr_spike_err)
        
        #230 chem blank value and error in fmol
        
        self.chem_blank_zero = (self.spike_nine_used * zero_nine_corr_spike) * 1000
        
        self.chem_blank_zero_err = abs(self.chem_blank_zero * zero_nine_corr_spike_err)
        '''
        while self.inquiry: 
            print "CHEMBLANK VALUES FOR AGE CALC: " 
            print "238U chemblank: " + str(self.chem_blank_eight) + " ± " + str(self.chem_blank_eight_err) + " pmol"
            print "232Th chemblank: " + str(self.chem_blank_two) + " ± " + str(self.chem_blank_two_err) + " pmol"
            print "230Th chemblank: " + str(self.chem_blank_zero) +  " ± " + str(self.chem_blank_zero_err) + " fmol"
            break
        '''    
        lstChemBlank = [self.chem_blank_eight,self.chem_blank_eight_err, self.chem_blank_two, self.chem_blank_two_err,
                        self.chem_blank_zero, self.chem_blank_zero_err]
        
        return lstChemBlank
    

            
            
            
     
        
        
        


        
root = tk.Tk()
app = Application(master=root)
app.mainloop()

